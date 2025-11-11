"""
Report Exporter Module
Handles PDF and Excel export functionality for security reports
"""

from datetime import datetime
from typing import Dict, List
import io

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.platypus import Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ReportExporter:
    """Export security reports to PDF and Excel formats"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a237e'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        
        # Subtitle style
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#283593'),
            spaceAfter=12,
            spaceBefore=12
        ))
        
        # Score style
        self.styles.add(ParagraphStyle(
            name='ScoreStyle',
            parent=self.styles['Normal'],
            fontSize=48,
            textColor=colors.HexColor('#2e7d32'),
            alignment=TA_CENTER,
            spaceAfter=20
        ))
    
    def export_to_pdf(self, report_data: Dict) -> bytes:
        """
        Export security report to PDF
        
        Args:
            report_data: Dictionary containing report data
            
        Returns:
            PDF file as bytes
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                               rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Title
        title = Paragraph(f"API Security Report", self.styles['CustomTitle'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # API Name
        api_name = Paragraph(f"<b>{report_data['api_name']}</b>", self.styles['Heading2'])
        elements.append(api_name)
        elements.append(Spacer(1, 12))
        
        # Date Range
        date_info = f"Report Period: {report_data['date_range']['start']} to {report_data['date_range']['end']}"
        elements.append(Paragraph(date_info, self.styles['Normal']))
        elements.append(Spacer(1, 6))
        
        # Generated Date
        gen_date = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(gen_date, self.styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Security Score
        score = report_data['score']['total_score']
        level = report_data['score']['security_level']
        
        score_color = self._get_score_color(score)
        score_text = f'<font color="{score_color}"><b>{score:.1f}/100</b></font>'
        elements.append(Paragraph(score_text, self.styles['ScoreStyle']))
        
        level_text = f'Security Level: <b>{level}</b>'
        elements.append(Paragraph(level_text, self.styles['Heading3']))
        elements.append(Spacer(1, 30))
        
        # Component Scores Table
        elements.append(Paragraph("Security Components", self.styles['CustomSubtitle']))
        elements.append(Spacer(1, 12))
        
        component_data = [['Component', 'Score', 'Weight', 'Contribution']]
        for component, comp_score in report_data['score']['component_scores'].items():
            weight = report_data['weights'].get(component, 0)
            contribution = comp_score * weight
            component_data.append([
                self._format_component_name(component),
                f"{comp_score:.1f}",
                f"{weight:.1%}",
                f"{contribution:.2f}"
            ])
        
        component_table = Table(component_data, colWidths=[3*inch, 1*inch, 1*inch, 1*inch])
        component_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(component_table)
        elements.append(Spacer(1, 30))
        
        # Recommendations
        if report_data['score']['recommendations']:
            elements.append(PageBreak())
            elements.append(Paragraph("Security Recommendations", self.styles['CustomSubtitle']))
            elements.append(Spacer(1, 12))
            
            for i, rec in enumerate(report_data['score']['recommendations'], 1):
                severity_color = self._get_severity_color(rec['severity'])
                rec_text = f"""
                <b>{i}. [{rec['severity'].upper()}] {rec['category'].title()}</b><br/>
                <font color="{severity_color}">● {rec['message']}</font><br/>
                <i>Action: {rec['action']}</i>
                """
                elements.append(Paragraph(rec_text, self.styles['Normal']))
                elements.append(Spacer(1, 12))
        
        # Traffic Statistics
        if 'traffic_stats' in report_data:
            elements.append(PageBreak())
            elements.append(Paragraph("Traffic Statistics", self.styles['CustomSubtitle']))
            elements.append(Spacer(1, 12))
            
            stats = report_data['traffic_stats']
            stats_data = [
                ['Metric', 'Value'],
                ['Total Requests', f"{stats.get('total_requests', 0):,}"],
                ['Unique IPs', f"{stats.get('unique_ips', 0):,}"],
                ['Avg Requests/Hour', f"{stats.get('avg_requests_per_hour', 0):.1f}"],
                ['Max Requests/Hour', f"{stats.get('max_requests_per_hour', 0):,}"],
                ['Error Rate', f"{stats.get('error_rate', 0):.2f}%"]
            ]
            
            stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(stats_table)
        
        # Build PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes
    
    def export_to_excel(self, report_data: Dict) -> bytes:
        """
        Export security report to Excel
        
        Args:
            report_data: Dictionary containing report data
            
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, report_data)
        self._create_components_sheet(wb, report_data)
        self._create_recommendations_sheet(wb, report_data)
        if 'traffic_stats' in report_data:
            self._create_traffic_sheet(wb, report_data)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        return excel_bytes
    
    def _create_summary_sheet(self, wb: Workbook, report_data: Dict):
        """Create summary sheet in Excel"""
        ws = wb.create_sheet("Summary", 0)
        
        # Header styling
        header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=14)
        
        # Title
        ws['A1'] = "API Security Report"
        ws['A1'].font = Font(bold=True, size=18, color="1a237e")
        ws.merge_cells('A1:D1')
        
        # API Info
        ws['A3'] = "API Name:"
        ws['B3'] = report_data['api_name']
        ws['A4'] = "Report Period:"
        ws['B4'] = f"{report_data['date_range']['start']} to {report_data['date_range']['end']}"
        ws['A5'] = "Generated:"
        ws['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Security Score
        ws['A7'] = "Security Score:"
        ws['B7'] = report_data['score']['total_score']
        ws['B7'].font = Font(bold=True, size=24, color=self._get_score_color_hex(report_data['score']['total_score']))
        
        ws['A8'] = "Security Level:"
        ws['B8'] = report_data['score']['security_level']
        ws['B8'].font = Font(bold=True, size=14)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
    
    def _create_components_sheet(self, wb: Workbook, report_data: Dict):
        """Create components sheet in Excel"""
        ws = wb.create_sheet("Components")
        
        # Headers
        headers = ['Component', 'Score', 'Weight', 'Contribution']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        row = 2
        for component, score in report_data['score']['component_scores'].items():
            weight = report_data['weights'].get(component, 0)
            contribution = score * weight
            
            ws.cell(row=row, column=1, value=self._format_component_name(component))
            ws.cell(row=row, column=2, value=round(score, 2))
            ws.cell(row=row, column=3, value=round(weight, 4))
            ws.cell(row=row, column=4, value=round(contribution, 2))
            row += 1
        
        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _create_recommendations_sheet(self, wb: Workbook, report_data: Dict):
        """Create recommendations sheet in Excel"""
        ws = wb.create_sheet("Recommendations")

        # Headers
        headers = ['#', 'Severity', 'Category', 'Message', 'Action']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data
        row = 2
        for i, rec in enumerate(report_data['score']['recommendations'], 1):
            ws.cell(row=row, column=1, value=i)

            severity_cell = ws.cell(row=row, column=2, value=rec['severity'].upper())
            severity_cell.font = Font(bold=True, color=self._get_severity_color_hex(rec['severity']))

            ws.cell(row=row, column=3, value=rec['category'].title())
            ws.cell(row=row, column=4, value=rec['message'])
            ws.cell(row=row, column=5, value=rec['action'])
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50

    def _create_traffic_sheet(self, wb: Workbook, report_data: Dict):
        """Create traffic statistics sheet in Excel"""
        ws = wb.create_sheet("Traffic Stats")

        # Headers
        ws['A1'] = "Metric"
        ws['B1'] = "Value"
        ws['A1'].font = Font(bold=True, color="FFFFFF")
        ws['B1'].font = Font(bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
        ws['B1'].fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")

        # Data
        stats = report_data['traffic_stats']
        metrics = [
            ('Total Requests', stats.get('total_requests', 0)),
            ('Unique IPs', stats.get('unique_ips', 0)),
            ('Avg Requests/Hour', round(stats.get('avg_requests_per_hour', 0), 2)),
            ('Max Requests/Hour', stats.get('max_requests_per_hour', 0)),
            ('Error Rate (%)', round(stats.get('error_rate', 0), 2))
        ]

        row = 2
        for metric, value in metrics:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Sensitive Data
        if 'sensitive_data' in stats and stats['sensitive_data'].get('has_sensitive_data'):
            ws.cell(row=row+1, column=1, value="Sensitive Data Found:")
            ws.cell(row=row+1, column=1).font = Font(bold=True, color="FF0000")
            row += 2

            ws.cell(row=row, column=1, value="Keyword")
            ws.cell(row=row, column=2, value="Count")
            ws.cell(row=row, column=3, value="Percentage")
            for col in range(1, 4):
                ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1

            for keyword, info in stats['sensitive_data']['sensitive_keywords'].items():
                ws.cell(row=row, column=1, value=keyword)
                ws.cell(row=row, column=2, value=info['count'])
                ws.cell(row=row, column=3, value=f"{info['percentage']:.2f}%")
                row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20

    def _format_component_name(self, component: str) -> str:
        """Format component name for display"""
        return component.replace('_', ' ').title()

    def _get_score_color(self, score: float) -> str:
        """Get color for score (for PDF)"""
        if score >= 90:
            return '#2e7d32'  # Green
        elif score >= 75:
            return '#558b2f'  # Light green
        elif score >= 60:
            return '#f9a825'  # Yellow
        elif score >= 40:
            return '#ef6c00'  # Orange
        else:
            return '#c62828'  # Red

    def _get_score_color_hex(self, score: float) -> str:
        """Get color hex for score (for Excel)"""
        color = self._get_score_color(score)
        return color.replace('#', '')

    def _get_severity_color(self, severity: str) -> str:
        """Get color for severity (for PDF)"""
        colors_map = {
            'critical': '#c62828',
            'high': '#ef6c00',
            'medium': '#f9a825',
            'low': '#558b2f'
        }
        return colors_map.get(severity.lower(), '#000000')

    def _get_severity_color_hex(self, severity: str) -> str:
        """Get color hex for severity (for Excel)"""
        color = self._get_severity_color(severity)
        return color.replace('#', '')

